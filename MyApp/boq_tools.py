"""Deterministic BOQ import and planning helpers.

The importer intentionally handles structured measurement schedules instead of
guessing construction intent from drawing geometry.  This keeps quantities
auditable and makes the workflow useful without an AI service.
"""

from __future__ import annotations

import csv
import io
import re
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook


MAX_IMPORT_ROWS = 500
HEADER_ALIASES = {
    'category': {'category', 'section', 'trade', 'work category', 'heading'},
    'description': {'description', 'item description', 'work description', 'item', 'scope', 'particulars'},
    'unit': {'unit', 'uom', 'unit of measure'},
    'quantity': {'quantity', 'qty', 'measured quantity'},
    'rate': {'rate', 'unit rate', 'price', 'unit price'},
    'amount': {'amount', 'total', 'total amount'},
}


class BOQImportError(ValueError):
    """A user-correctable BOQ import error."""


def _text(value):
    return '' if value is None else str(value).strip()


def _header(value):
    return re.sub(r'[^a-z0-9]+', ' ', _text(value).lower()).strip()


def _decimal(value, field, row_number, *, optional=False):
    text = _text(value).replace(',', '')
    if not text and optional:
        return ''
    try:
        number = Decimal(text)
    except InvalidOperation as exc:
        raise BOQImportError(f'Row {row_number}: {field} must be a valid number.') from exc
    if number < 0:
        raise BOQImportError(f'Row {row_number}: {field} cannot be negative.')
    return format(number.normalize(), 'f')


def _column_map(row):
    result = {}
    for index, value in enumerate(row):
        normalized = _header(value)
        for canonical, aliases in HEADER_ALIASES.items():
            if normalized in aliases and canonical not in result:
                result[canonical] = index
    return result


def _parse_rows(raw_rows):
    rows = list(raw_rows)
    header_index = None
    columns = None
    for index, row in enumerate(rows[:25]):
        candidate = _column_map(row)
        if 'description' in candidate and 'quantity' in candidate:
            header_index, columns = index, candidate
            break
    if columns is None:
        raise BOQImportError(
            'No BOQ header was found. Include at least Description and Quantity columns.'
        )

    parsed = []
    current_category = 'General'
    for offset, row in enumerate(rows[header_index + 1:], start=header_index + 2):
        def get(name):
            return row[columns[name]] if name in columns and columns[name] < len(row) else ''

        description = _text(get('description'))
        category = _text(get('category')) or current_category
        quantity_value = _text(get('quantity'))
        if not description and not quantity_value:
            continue
        if not description:
            raise BOQImportError(f'Row {offset}: description is required.')
        # A heading row may omit quantity. Remember it for the following items.
        if not quantity_value:
            current_category = description[:40]
            continue
        current_category = category
        quantity = _decimal(quantity_value, 'quantity', offset)
        rate = _decimal(get('rate'), 'rate', offset, optional=True)
        amount = _decimal(get('amount'), 'amount', offset, optional=True)
        if not amount and rate:
            amount = format((Decimal(quantity) * Decimal(rate)).normalize(), 'f')
        parsed.append({
            'category': category[:40],
            'description': description[:500],
            'unit': _text(get('unit'))[:40],
            'quantity': quantity,
            'rate': rate or '0',
            'amount': amount or '0',
        })
        if len(parsed) > MAX_IMPORT_ROWS:
            raise BOQImportError(f'BOQ files are limited to {MAX_IMPORT_ROWS} item rows per import.')
    if not parsed:
        raise BOQImportError('The uploaded file does not contain any measurable BOQ item rows.')
    return parsed


def parse_boq_upload(upload):
    """Parse a structured XLSX or CSV upload into normalized BOQ rows."""
    filename = upload.name.lower()
    if filename.endswith('.xlsx'):
        try:
            workbook = load_workbook(upload, read_only=True, data_only=True)
            sheet = workbook.active
            return _parse_rows(sheet.iter_rows(values_only=True))
        except BOQImportError:
            raise
        except Exception as exc:
            raise BOQImportError('The Excel workbook could not be read. Check that it is a valid .xlsx file.') from exc
        finally:
            if 'workbook' in locals():
                workbook.close()
    if filename.endswith('.csv'):
        raw_content = upload.read()
        try:
            content = raw_content.decode('utf-8-sig')
        except UnicodeDecodeError:
            content = raw_content.decode('cp1252')
        return _parse_rows(csv.reader(io.StringIO(content)))
    raise BOQImportError('Upload an Excel (.xlsx) or CSV (.csv) BOQ/measurement schedule.')


def normalized_tokens(value):
    ignored = {'and', 'for', 'with', 'the', 'work', 'works', 'supply', 'installation', 'install'}
    return {
        token for token in re.findall(r'[a-z0-9]+', value.lower())
        if len(token) > 2 and token not in ignored
    }


def suggest_material(description, materials):
    """Return the best deterministic material-name match, if credible."""
    target = normalized_tokens(description)
    best, best_score = None, 0
    for candidate in materials:
        candidate_tokens = normalized_tokens(candidate.name)
        if not candidate_tokens:
            continue
        overlap = len(target & candidate_tokens)
        score = overlap / len(candidate_tokens)
        if candidate.name.lower() in description.lower():
            score += 1
        if score > best_score:
            best, best_score = candidate, score
    return best if best_score >= 0.5 else None


METHOD_RULES = (
    (('demolition', 'removal', 'dismantl'), 'Demolition and removal',
     'Barricade the area, identify live services, remove items in a controlled sequence, segregate debris and dispose through approved routes.'),
    (('electrical', 'cable', 'lighting', 'socket'), 'Electrical installation',
     'Verify approved shop drawings, isolate supplies, install containment and cables, terminate with identification, then perform continuity and insulation-resistance tests.'),
    (('hvac', 'duct', 'air condition'), 'HVAC installation',
     'Coordinate openings and services, install supports and equipment to approved drawings, seal joints, pressure-test where applicable, balance the system and record results.'),
    (('fire alarm', 'fire fighting', 'sprinkler'), 'Fire and life-safety systems',
     'Use authority-approved materials, coordinate zones and interfaces, install and label devices, complete pressure/functional testing and retain commissioning records.'),
    (('gypsum', 'partition', 'ceiling'), 'Partitions and ceilings',
     'Set out from approved drawings, install tracks and supports, coordinate concealed services, fix boards/panels, treat joints and verify line, level and finish.'),
    (('paint', 'coating'), 'Painting and coatings',
     'Approve samples, protect adjacent finishes, prepare and clean substrates, apply primer and specified coats, observe curing times and inspect colour and coverage.'),
    (('floor', 'tile', 'carpet', 'vinyl'), 'Floor finishes',
     'Confirm substrate moisture and level, approve setting-out, apply specified adhesive/system, maintain joints and levels, then protect completed finishes.'),
    (('joinery', 'cabinet', 'door', 'wood'), 'Joinery installation',
     'Verify site dimensions and approved samples, fabricate to shop drawings, protect finished surfaces, install plumb and secure, adjust hardware and complete snag rectification.'),
)


def method_statement_for(category, descriptions):
    combined = f"{category} {' '.join(descriptions)}".lower()
    title = category or 'General works'
    procedure = (
        'Confirm approved drawings, specifications, material approvals and site readiness. '
        'Set out the work, execute using approved materials and competent personnel, inspect each stage and close all recorded snags.'
    )
    for keywords, matched_title, matched_procedure in METHOD_RULES:
        if any(keyword in combined for keyword in keywords):
            title, procedure = matched_title, matched_procedure
            break
    return {
        'title': title[:100],
        'scope': '; '.join(descriptions),
        'preparation': 'Review approved drawings and specifications; verify dimensions and interfaces; secure permits, access, tools, materials and approved samples before starting.',
        'procedure': procedure,
        'quality': 'Use approved inspection checklists. Verify materials on receipt, inspect concealed work before closure, record test results and obtain the required stage approvals.',
        'safety': 'Complete task risk assessment and toolbox talk. Use required PPE, maintain access and housekeeping, isolate live services and follow site permit and emergency procedures.',
        'completion': 'Clean and protect completed work, complete testing and snag rectification, submit inspection records and hand over the finished area for acceptance.',
    }


def build_work_packages(rows):
    groups = {}
    for row in rows:
        groups.setdefault(row['category'] or 'General', []).append(row['description'])
    return [method_statement_for(category, descriptions) for category, descriptions in groups.items()]
