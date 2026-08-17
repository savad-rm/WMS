from copy import copy
from html import escape
from io import BytesIO
from math import ceil
from pathlib import Path
from types import SimpleNamespace

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.drawing.image import Image as ExcelImage
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Flowable, Image, KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

from .quotation_document import (
    DEFAULT_INTRODUCTION, plain_rich_text, presentation_rows, rich_text_to_html,
    unpack_document,
)


ASSET_DIR = Path(settings.BASE_DIR) / 'MyApp' / 'quotation_templates'
EXCEL_TEMPLATE = ASSET_DIR / 'exalter_quotation_template.xlsx'
DETAILED_EXCEL_TEMPLATE = ASSET_DIR / 'quotation_sample.xlsx'
LETTERHEAD_IMAGE = ASSET_DIR / 'exalter_letterhead.png'
SIGNATURE_IMAGE = ASSET_DIR / 'exalter_signature.png'
STAMP_IMAGE = ASSET_DIR / 'exalter_stamp.png'


class _PaginatedQuotationTable(Flowable):
    """Split quotation rows at A4 boundaries without splitting a row.

    ReportLab cannot split a table across a page when it contains a vertical
    cell span.  Quotations use those spans for lump-sum rate and amount cells,
    so this flowable creates a fresh, repeated-header table for each page
    fragment.  A lump-sum span is rebuilt inside each fragment rather than
    forcing its complete group onto the next page.
    """

    def __init__(self, build_fragment, row_count, start=0):
        super().__init__()
        self.build_fragment = build_fragment
        self.row_count = row_count
        self.start = start

    def wrap(self, available_width, available_height):
        # Ask Platypus to call split(), where the current page's actual usable
        # height is available and a correctly sized fragment can be produced.
        return available_width, available_height + 1

    def split(self, available_width, available_height):
        for end in range(self.row_count, self.start, -1):
            fragment = self.build_fragment(self.start, end)
            _, height = fragment.wrap(available_width, available_height)
            if height <= available_height:
                flowables = [fragment]
                if end < self.row_count:
                    flowables.append(
                        _PaginatedQuotationTable(
                            self.build_fragment, self.row_count, start=end,
                        )
                    )
                return flowables
        # There is no room for even one complete row in the current frame;
        # Platypus will move this continuation to the next page.
        return []

def quotation_amount_words(value):
    ones = (
        'zero', 'one', 'two', 'three', 'four', 'five', 'six', 'seven', 'eight',
        'nine', 'ten', 'eleven', 'twelve', 'thirteen', 'fourteen', 'fifteen',
        'sixteen', 'seventeen', 'eighteen', 'nineteen',
    )
    tens = ('', '', 'twenty', 'thirty', 'forty', 'fifty', 'sixty', 'seventy', 'eighty', 'ninety')

    def integer_words(number):
        if number < 20:
            return ones[number]
        if number < 100:
            return tens[number // 10] + (f'-{ones[number % 10]}' if number % 10 else '')
        if number < 1_000:
            return f'{ones[number // 100]} hundred' + (
                f' and {integer_words(number % 100)}' if number % 100 else ''
            )
        for size, label in ((1_000_000_000, 'billion'), (1_000_000, 'million'), (1_000, 'thousand')):
            if number >= size:
                return f'{integer_words(number // size)} {label}' + (
                    f' {integer_words(number % size)}' if number % size else ''
                )
        return str(number)

    amount = value.quantize(value.__class__('0.01'))
    riyals = int(amount)
    dirhams = int((amount - riyals) * 100)
    result = integer_words(riyals)
    if dirhams:
        result += f' and {integer_words(dirhams)} dirhams'
    return result


def _client_nameplate(value):
    return value if value.strip().lower().startswith('m/s') else f'M/s {value}'


def _copy_row_style(sheet, source_row, target_row):
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(2, 8):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def _quotation_lines(quote):
    lines_source = quote.lines
    lines = list(lines_source.all()) if hasattr(lines_source, 'all') else list(lines_source)
    if lines:
        return lines
    return [SimpleNamespace(
        item_code='1', description=quote.details or 'Quotation total', unit='lot',
        quantity=quote.amount.__class__('1'), unit_rate=quote.amount,
        amount=quote.amount,
    )]


def _insert_line_rows(sheet, count):
    extra = max(0, count - 1)
    if not extra:
        return 0
    shifted_merges = []
    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row >= 19:
            shifted_merges.append((str(merged), merged.min_col, merged.min_row, merged.max_col, merged.max_row))
            sheet.unmerge_cells(str(merged))
    sheet.insert_rows(19, extra)
    for _, min_col, min_row, max_col, max_row in shifted_merges:
        sheet.merge_cells(
            start_row=min_row + extra, start_column=min_col,
            end_row=max_row + extra, end_column=max_col,
        )
    for row in range(19, 19 + extra):
        _copy_row_style(sheet, 18, row)
    return extra


def build_quotation_excel(quote):
    workbook = load_workbook(DETAILED_EXCEL_TEMPLATE)
    sheet = workbook['Quote']
    for extra_sheet in list(workbook.worksheets):
        if extra_sheet != sheet:
            workbook.remove(extra_sheet)
    enquiry = quote.ENQUIRY
    document = unpack_document(quote.details, quote.validity_days)
    client = document.get('client') or {}
    client_name = client.get('name', enquiry.client_name)
    client_phone = client.get('phone', enquiry.client_phone)
    client_email = client.get('email', enquiry.client_email)
    rows = presentation_rows(_quotation_lines(quote))

    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row >= 19:
            sheet.unmerge_cells(str(merged))
    sheet.delete_rows(19, sheet.max_row - 18)

    sheet['B9'] = _client_nameplate(client_name)
    sheet['B9'].font = Font(name='Helvetica', size=10, bold=True)
    sheet['B10'] = f'Phone: {client_phone}' if client_phone else 'Phone:'
    sheet['B11'] = f'Email: {client_email}' if client_email else 'Email:'
    sheet['B12'] = quote.client_address or 'Doha - State of Qatar'
    sheet['B12'].alignment = Alignment(wrap_text=True, vertical='top')
    sheet['F9'] = quote.display_number
    sheet['F10'] = quote.issue_date
    sheet['F10'].number_format = 'dd-mmm-yy'
    sheet['C15'] = f'Subject : {quote.subject}'
    sheet['B16'] = 'Dear Sir,'
    sheet['C17'] = quote.introduction or DEFAULT_INTRODUCTION

    thin = Side(style='thin', color='000000')
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    body_font = Font(name='Helvetica', size=10)
    bold_font = Font(name='Helvetica', size=10, bold=True)
    money_font = Font(name='Helvetica', size=11)
    current_row = 19
    section_item_rows = []
    section_total_rows = []
    all_item_rows = []
    lump_sum_ranges = {}

    def sum_formula(row_numbers):
        return '=' + '+'.join(f'G{number}' for number in row_numbers) if row_numbers else '=0'

    for entry_index, entry in enumerate(rows):
        kind = entry['kind']
        if kind in ('section', 'subheading', 'note'):
            line = entry['line']
            group_start = entry.get('lump_sum_group_start')
            has_group_children = bool(group_start and any(
                later.get('lump_sum_group') == group_start
                for later in rows[entry_index + 1:]
            ))
            sheet.cell(
                current_row, 2,
                line.item_code if kind in ('section', 'subheading') else '',
            )
            sheet.cell(current_row, 3, plain_rich_text(line.description))
            heading_end_column = (
                6 if kind in ('section', 'subheading') else 7
            )
            sheet.merge_cells(
                start_row=current_row, start_column=3,
                end_row=current_row, end_column=heading_end_column,
            )
            sheet.cell(current_row, 3).font = bold_font if kind != 'note' else body_font
            sheet.cell(current_row, 3).alignment = Alignment(
                horizontal='center' if kind == 'section' else 'left',
                vertical='center', wrap_text=True,
            )
            if kind in ('section', 'subheading') and line.amount and not has_group_children:
                sheet.cell(current_row, 7, float(line.amount))
                sheet.cell(current_row, 7).number_format = '#,##0.00'
                sheet.cell(current_row, 7).font = money_font
                section_item_rows.append(current_row)
        elif kind == 'item':
            line = entry['line']
            values = (
                line.item_code, plain_rich_text(line.description), line.unit,
                float(line.quantity) if line.quantity else None,
                float(line.unit_rate) if line.unit_rate else None,
            )
            for column, value in enumerate(values, start=2):
                sheet.cell(current_row, column, value)
            sheet.cell(
                current_row, 7,
                f'=E{current_row}*F{current_row}' if line.quantity and line.unit_rate else None,
            )
            for column in (5, 6, 7):
                sheet.cell(current_row, column).number_format = '#,##0.00'
            sheet.cell(current_row, 6).font = money_font
            sheet.cell(current_row, 7).font = money_font
            sheet.cell(current_row, 3).alignment = Alignment(
                horizontal='left', wrap_text=True, vertical='center',
            )
            section_item_rows.append(current_row)
            all_item_rows.append(current_row)
            if entry.get('lump_sum_group'):
                group_id, group_total = entry['lump_sum_group']
                group = lump_sum_ranges.setdefault(group_id, {'start': current_row, 'end': current_row, 'total': group_total})
                group['end'] = current_row
        else:
            sheet.cell(current_row, 2, entry.get('code', ''))
            sheet.cell(current_row, 3, entry['label'])
            sheet.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
            sheet.cell(current_row, 4, 'Sub-Total (QAR)')
            sheet.cell(current_row, 7, float(entry['amount']))
            sheet.cell(current_row, 7).number_format = '#,##0.00'
            for column in range(2, 8):
                sheet.cell(current_row, column).font = bold_font if column != 7 else Font(name='Helvetica', size=11, bold=True)
            section_item_rows = []
            section_total_rows.append(current_row)
        for column in range(2, 8):
            cell = sheet.cell(current_row, column)
            cell.border = grid
            if cell.font.name != 'Helvetica':
                cell.font = body_font
            if column in (2, 4, 5):
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if kind == 'item':
            description_length = len(plain_rich_text(entry['line'].description))
            sheet.row_dimensions[current_row].height = max(28, 14 * ceil(description_length / 48))
        else:
            sheet.row_dimensions[current_row].height = 22
        current_row += 1

    for group in lump_sum_ranges.values():
        for column in (6, 7):
            if group['end'] > group['start']:
                sheet.merge_cells(
                    start_row=group['start'], start_column=column,
                    end_row=group['end'], end_column=column,
                )
            cell = sheet.cell(group['start'], column)
            cell.value = float(group['total'])
            cell.number_format = '#,##0.00'
            cell.font = money_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
    sheet.cell(
        current_row, 2,
        f'Grand Total (QAR - {quotation_amount_words(quote.amount).capitalize()} only)',
    )
    sheet.cell(current_row, 7, float(quote.amount))
    for column in range(2, 8):
        sheet.cell(current_row, column).border = grid
        sheet.cell(current_row, column).font = bold_font if column != 7 else Font(name='Helvetica', size=11, bold=True)
    sheet.cell(current_row, 7).number_format = '#,##0.00'
    current_row += 1
    sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=7)
    sheet.cell(current_row, 2, 'Specification/Clarification')
    sheet.cell(current_row, 2).font = Font(name='Helvetica', size=10, bold=True, underline='single')
    current_row += 2
    for index, term in enumerate(document['terms'], start=1):
        sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=7)
        sheet.cell(current_row, 2, f'{index}. {term["title"]}')
        sheet.cell(current_row, 2).font = Font(name='Helvetica', size=10, bold=True, underline='single')
        current_row += 1
        for text in term['body'].splitlines() or ['']:
            sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=7)
            sheet.cell(current_row, 2, text)
            sheet.cell(current_row, 2).font = body_font
            sheet.cell(current_row, 2).alignment = Alignment(wrap_text=True, vertical='top')
            current_row += 1
        current_row += 1
    for value, bold in (
        (quote.closing_text, False), ('With Best Regards,', True),
        ('Exalter Trading & Contracting', True), (quote.signatory_name, True),
        (quote.signatory_title, True),
        (f'Mobile {quote.signatory_phone}' if quote.signatory_phone else '', True),
    ):
        sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=7)
        sheet.cell(current_row, 2, value)
        sheet.cell(current_row, 2).font = bold_font if bold else body_font
        sheet.cell(current_row, 2).alignment = Alignment(wrap_text=True)
        current_row += 1
    if SIGNATURE_IMAGE.exists() and STAMP_IMAGE.exists():
        signature = ExcelImage(str(SIGNATURE_IMAGE))
        signature.width, signature.height = 135, 54
        stamp = ExcelImage(str(STAMP_IMAGE))
        stamp.width, stamp.height = 82, 82
        sheet.add_image(signature, f'C{max(19, current_row - 4)}')
        sheet.add_image(stamp, f'E{max(19, current_row - 5)}')
        current_row += 2
    sheet.print_area = f'B9:G{current_row}'
    sheet['F10'].alignment = Alignment(horizontal='center', vertical='center')
    for column in range(2, 8):
        sheet.cell(18, column).alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True,
        )
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_setup.orientation = 'portrait'
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = '18:18'
    sheet.oddFooter.center.text = ''
    sheet.evenFooter.center.text = ''

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_quotation_pdf(quote):
    output = BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4, leftMargin=22 * mm, rightMargin=22 * mm,
        topMargin=34 * mm, bottomMargin=26 * mm,
        title=quote.display_number, author='Exalter Trading & Contracting',
    )

    def draw_letterhead(canvas, document):
        canvas.saveState()
        canvas.drawImage(
            str(LETTERHEAD_IMAGE), 0, 0, width=A4[0], height=A4[1],
            preserveAspectRatio=False, mask='auto',
        )
        canvas.setFillColor(colors.HexColor('#8c7428'))
        canvas.setFont('Helvetica', 7)
        canvas.drawRightString(A4[0] - 14 * mm, 13 * mm, f'Page {document.page}')
        canvas.restoreState()

    styles = getSampleStyleSheet()
    normal = ParagraphStyle('QuotationBody', parent=styles['BodyText'], fontName='Helvetica', fontSize=10, leading=12)
    centered = ParagraphStyle('QuotationCentered', parent=normal, alignment=TA_CENTER)
    money = ParagraphStyle('QuotationMoney', parent=normal, fontSize=11, alignment=TA_CENTER)
    heading = ParagraphStyle('QuotationHeading', parent=normal, fontName='Helvetica-Bold', spaceBefore=7, spaceAfter=2)
    section_heading = ParagraphStyle(
        'QuotationSection', parent=normal, fontName='Helvetica-Bold', alignment=TA_CENTER,
    )
    subheading_heading = ParagraphStyle(
        'QuotationSubheading', parent=normal, fontName='Helvetica-Bold', alignment=0,
    )
    subject = ParagraphStyle('QuotationSubject', parent=normal, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceBefore=10, spaceAfter=6)

    record = quote.ENQUIRY
    document = unpack_document(quote.details, quote.validity_days)
    client = document.get('client') or {}
    client_name = client.get('name', record.client_name)
    client_phone = client.get('phone', record.client_phone)
    client_email = client.get('email', record.client_email)
    client_lines = [f'<b>{escape(_client_nameplate(client_name))}</b>']
    if client_phone:
        client_lines.append(f'Phone: {escape(client_phone)}')
    if client_email:
        client_lines.append(f'Email: {escape(client_email)}')
    if quote.client_address:
        client_lines.append(escape(quote.client_address).replace('\n', '<br/>'))
    reference = Table([
        [Paragraph(f'<b>{quote.display_number}</b>', ParagraphStyle('Ref', parent=normal, alignment=2))],
        [Paragraph(f'<b>{quote.issue_date:%d-%b-%y}</b>', ParagraphStyle('Date', parent=normal, alignment=TA_CENTER))],
        [Paragraph('<b>QUOTATION</b>', ParagraphStyle('Plate', parent=normal, alignment=TA_CENTER))],
    ], colWidths=[48 * mm], rowHeights=[None, None, 6 * mm])
    reference.setStyle(TableStyle([('BOX', (0, 2), (0, 2), 1, colors.black)]))
    info = Table([[Paragraph('<br/>'.join(client_lines), normal), reference]], colWidths=[105 * mm, 48 * mm])
    info.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP'), ('ALIGN', (1, 0), (1, 0), 'RIGHT')]))

    story = [info]
    story.extend([
        Paragraph(f'Subject : <u>{quote.subject}</u>', subject),
        Paragraph('<b>Dear Sir,</b>', normal),
        Paragraph(f'&nbsp;&nbsp;&nbsp;&nbsp;{quote.introduction or DEFAULT_INTRODUCTION}', normal),
        Spacer(1, 3 * mm),
    ])

    table_data = [[
        Paragraph('<b>ITEMS</b>', centered), Paragraph('<b>DESCRIPTION</b>', centered),
        Paragraph('<b>UNIT</b>', centered), Paragraph('<b>QTY</b>', centered),
        Paragraph('<b>RATE</b>', money), Paragraph('<b>AMOUNT</b>', money),
    ]]
    row_spans = []
    bold_rows = []
    lump_sum_row_ranges = {}
    quotation_entries = presentation_rows(_quotation_lines(quote))
    for entry_index, entry in enumerate(quotation_entries):
        kind = entry['kind']
        row_index = len(table_data)
        if kind in ('section', 'subheading', 'note'):
            line = entry['line']
            group_start = entry.get('lump_sum_group_start')
            has_group_children = bool(group_start and any(
                later.get('lump_sum_group') == group_start
                for later in quotation_entries[entry_index + 1:]
            ))
            line_style = (
                section_heading if kind == 'section'
                else subheading_heading if kind == 'subheading'
                else normal
            )
            table_data.append([
                line.item_code if kind in ('section', 'subheading') else '',
                Paragraph(rich_text_to_html(line.description), line_style), '', '', '',
                Paragraph(f'{line.amount:,.2f}', money) if line.amount and not has_group_children else '',
            ])
            row_spans.append((
                row_index, 1,
                4 if kind in ('section', 'subheading') else 5,
            ))
            if kind != 'note':
                bold_rows.append(row_index)
        elif kind in ('subtotal', 'section_total'):
            table_data.append([
                entry.get('code', ''), Paragraph(f'<b>{entry["label"]}</b>', normal),
                Paragraph('<b>Sub-Total (QAR)</b>', normal), '', '',
                Paragraph(f'<b>{entry["amount"]:,.2f}</b>', money),
            ])
            row_spans.append((row_index, 2, 4))
            bold_rows.append(row_index)
        else:
            line = entry['line']
            table_data.append([
                line.item_code, Paragraph(rich_text_to_html(line.description), normal), line.unit,
                f'{line.quantity:,.2f}' if line.quantity else '',
                Paragraph(f'{line.unit_rate:,.2f}', money) if line.unit_rate else '',
                Paragraph(f'{line.amount:,.2f}', money) if line.amount else '',
            ])
            if entry.get('lump_sum_group'):
                group_id, group_total = entry['lump_sum_group']
                group = lump_sum_row_ranges.setdefault(
                    group_id, {'start': row_index, 'end': row_index, 'total': group_total},
                )
                group['end'] = row_index
    # Populate the lump-sum cells before constructing the ReportLab table.
    # Table copies its input data, so assigning these values afterwards leaves
    # the rendered (vertically merged) cells blank.
    for group in lump_sum_row_ranges.values():
        table_data[group['start']][4] = Paragraph(f'{group["total"]:,.2f}', money)
        table_data[group['start']][5] = Paragraph(f'{group["total"]:,.2f}', money)
    column_widths = [12 * mm, 66 * mm, 14 * mm, 12 * mm, 23 * mm, 27 * mm]
    base_table_style = [
        ('GRID', (0, 0), (-1, -1), .55, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (1, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3), ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]

    def build_line_fragment(start_index, end_index):
        """Build one page-safe table fragment with its own header and spans."""
        fragment_data = [list(table_data[0])]
        fragment_data.extend(list(row) for row in table_data[start_index + 1:end_index + 1])
        fragment_style = list(base_table_style)
        first_row = start_index + 1
        last_row = end_index

        for row_index, start_column, end_column in row_spans:
            if first_row <= row_index <= last_row:
                local_row = row_index - start_index
                fragment_style.append(
                    ('SPAN', (start_column, local_row), (end_column, local_row))
                )
        for row_index in bold_rows:
            if first_row <= row_index <= last_row:
                local_row = row_index - start_index
                fragment_style.append(
                    ('FONTNAME', (0, local_row), (-1, local_row), 'Helvetica-Bold')
                )
        for row_index, entry in enumerate(quotation_entries, start=1):
            if not first_row <= row_index <= last_row:
                continue
            local_row = row_index - start_index
            if entry['kind'] == 'section':
                fragment_style.extend([
                    ('ALIGN', (1, local_row), (4, local_row), 'CENTER'),
                    ('TOPPADDING', (0, local_row), (-1, local_row), 2),
                    ('BOTTOMPADDING', (0, local_row), (-1, local_row), 2),
                ])
            elif entry['kind'] == 'subheading':
                fragment_style.extend([
                    ('ALIGN', (1, local_row), (4, local_row), 'LEFT'),
                    ('TOPPADDING', (0, local_row), (-1, local_row), 1),
                    ('BOTTOMPADDING', (0, local_row), (-1, local_row), 1),
                ])

        for group in lump_sum_row_ranges.values():
            group_start = group['start']
            group_end = group['end']
            if group_end < first_row or group_start > last_row:
                continue
            local_start = max(group_start, first_row) - start_index
            local_end = min(group_end, last_row) - start_index
            if group_end > group_start:
                # A cross-page lump sum is visually merged within each page
                # fragment.  The amount is printed only in the final fragment
                # so it is never duplicated in the quotation total.
                for column in (4, 5):
                    fragment_style.append(
                        ('SPAN', (column, local_start), (column, local_end))
                    )
                    fragment_data[local_start][column] = (
                        Paragraph(f'{group["total"]:,.2f}', money)
                        if group_end <= last_row else ''
                    )

        fragment = Table(
            fragment_data, colWidths=column_widths, splitByRow=1, splitInRow=0,
        )
        fragment.setStyle(TableStyle(fragment_style))
        return fragment

    line_table = _PaginatedQuotationTable(
        build_line_fragment, len(table_data) - 1,
    )
    grand_total_style = ParagraphStyle(
        'QuotationGrandTotal', parent=normal, fontName='Helvetica-Bold',
        fontSize=10, leading=11, alignment=TA_CENTER,
    )
    grand_total_table = Table([[
        Paragraph(
            f'Grand Total (QAR - {quotation_amount_words(quote.amount).capitalize()} only)',
            grand_total_style,
        ),
        '', '', '', '', Paragraph(f'<b>{quote.amount:,.2f}</b>', money),
    ]], colWidths=[12 * mm, 66 * mm, 14 * mm, 12 * mm, 23 * mm, 27 * mm])
    grand_total_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), .55, colors.black),
        ('SPAN', (0, 0), (4, 0)),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (4, 0), 'CENTER'),
        ('ALIGN', (5, 0), (5, 0), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    # Keep the total outside the repeating line-item table. If it genuinely
    # needs a continuation page, it appears as a standalone total without a
    # misleading repeated ITEMS/DESCRIPTION header.
    story.extend([
        line_table, grand_total_table,
        Paragraph('<u><b>Specification/Clarification</b></u>', heading),
    ])

    for index, term in enumerate(document['terms'], start=1):
        term_flowables = [Paragraph(f'<u>{index}. {term["title"]}</u>', heading)]
        term_flowables.extend(
            Paragraph(line.strip(), normal)
            for line in term['body'].splitlines() if line.strip()
        )
        story.append(KeepTogether(term_flowables))

    signatory = [
        Paragraph('<b>With Best Regards,</b>', normal),
        Paragraph('<b>Exalter Trading &amp; Contracting</b>', normal),
    ]
    if SIGNATURE_IMAGE.exists():
        signatory.append(Image(str(SIGNATURE_IMAGE), width=40 * mm, height=16 * mm))
    signatory.extend([
        Paragraph(f'<b>{quote.signatory_name}</b>', normal),
        Paragraph(f'<b>{quote.signatory_title}</b>', normal),
        Paragraph(f'<b>Mobile {quote.signatory_phone}</b>', normal),
    ])
    approval_visual = Image(str(STAMP_IMAGE), width=27 * mm, height=27 * mm) if STAMP_IMAGE.exists() else ''
    signature_table = Table([[signatory, approval_visual]], colWidths=[112 * mm, 40 * mm])
    signature_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0), ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    story.extend([
        Spacer(1, 4 * mm), Paragraph(quote.closing_text, normal), Spacer(1, 4 * mm),
        KeepTogether(signature_table),
    ])
    doc.build(story, onFirstPage=draw_letterhead, onLaterPages=draw_letterhead)
    output.seek(0)
    return output
