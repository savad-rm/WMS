import json
import sqlite3
import tarfile
import tempfile
from unittest.mock import patch
from contextlib import closing
from decimal import Decimal
from io import BytesIO
from datetime import timedelta
from pathlib import Path

from django.db import IntegrityError, transaction
from django.contrib.auth.hashers import check_password, make_password
from django.core import mail
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import Client, TestCase, TransactionTestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from .models import (
    chat, enquiry, enquiry_attachment, estimate, login, material, material_request,
    material_required, project, project_document, project_manager_allocation, quotation,
    schedule, staff, supervisor_allocation, work, work_progress, workflow_notification,
)
from .deadline_notifications import ensure_quotation_deadline_notifications
from .quotation_document import (
    default_terms, pack_document, presentation_rows, quotation_tracking, unpack_document,
)
from .quotation_exports import build_quotation_excel, quotation_amount_words


@override_settings(
    EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
    DEFAULT_FROM_EMAIL='quotations@exalter.example',
)
class WorkflowTests(TestCase):
    def create_user(self, role, index):
        account = login.objects.create(
            username=f'user{index}@example.com', password='test-password', usertype=role
        )
        person = staff.objects.create(
            LOGIN=account, name=role, dob='', phone=f'90000000{index:02d}',
            email=account.username, photo='', place='', nation='', phone2='', designation=role,
        )
        return account, person

    def sign_in_as(self, account, person):
        session = self.client.session
        session['lid'] = account.pk
        session['sid'] = person.pk
        session.save()

    def setUp(self):
        self.executive = self.create_user('Marketing Executive', 1)
        self.manager = self.create_user('Marketing Manager', 2)
        self.estimator = self.create_user('Estimator', 3)
        self.accountant = self.create_user('Accountant', 4)
        self.project_manager = self.create_user('Project Manager', 5)
        self.controller = self.create_user('Document Controller', 6)

    def test_staff_email_is_unique(self):
        with self.assertRaises(IntegrityError), transaction.atomic():
            login.objects.create(username=self.executive[0].username, password='x', usertype='Estimator')

    def test_admin_project_transfer_locks_awarded_enquiry_and_uses_its_client_details(self):
        admin = self.create_user('Admin', 7)
        record = enquiry.objects.create(
            title='Awarded fit-out', client_name='Awarded Client',
            client_email='awarded@example.com', client_phone='5551234',
            created_by=self.executive[0], status='awarded',
        )
        self.sign_in_as(*admin)
        response = self.client.post(reverse('Add_project_post'), {
            'project_no': 'P-AWARD-1', 't1': 'Awarded Fit-out Project',
            'client_name': 'Incorrect Form Client', 'phone': '0000000',
            'email': 'incorrect@example.com', 'place': 'Doha', 'unit_no': '',
            'project_value': '1000', 'starting_date': '2026-08-01',
            'handout_date': '2026-09-01', 'project_duration': '31 Days',
            'project_area': '', 'project_type': '', 'textfield13': '',
            'enquiry': record.pk,
        })
        self.assertEqual(response.status_code, 302)
        created = project.objects.get(project_no='P-AWARD-1')
        record.refresh_from_db()
        self.assertEqual(record.PROJECT_id, created.pk)
        self.assertEqual(created.client_name, record.client_name)
        self.assertEqual(created.email, record.client_email)

        self.client.post(reverse('Add_project_post'), {
            'project_no': 'P-AWARD-2', 't1': 'Duplicate Project',
            'client_name': 'Client', 'phone': '0000000', 'email': 'x@example.com',
            'place': 'Doha', 'unit_no': '', 'project_value': '1000',
            'starting_date': '2026-08-01', 'handout_date': '2026-09-01',
            'project_duration': '31 Days', 'project_area': '', 'project_type': '',
            'textfield13': '', 'enquiry': record.pk,
        })
        self.assertFalse(project.objects.filter(project_no='P-AWARD-2').exists())

    def test_legacy_password_is_upgraded_after_login(self):
        response = self.client.post(reverse('login_post'), {
            'username': self.executive[0].username, 'password': 'test-password',
        })
        self.assertRedirects(response, reverse('workflow_dashboard'))
        self.executive[0].refresh_from_db()
        self.assertTrue(check_password('test-password', self.executive[0].password))

    def test_complete_quotation_approval_and_award_flow(self):
        self.sign_in_as(*self.executive)
        response = self.client.post(reverse('workflow_add_enquiry'), {
            'title': 'Office fit-out', 'client_name': 'Example Client',
            'client_email': 'client@example.com', 'description': 'Fit-out scope',
            'quotation_deadline': (timezone.now() + timedelta(days=10)).isoformat(),
        })
        record = enquiry.objects.get()
        self.assertRedirects(response, reverse('workflow_detail', args=(record.pk,)))

        self.sign_in_as(*self.manager)
        response = self.client.post(reverse('workflow_assign', args=(record.pk,)), {
            'estimator': self.estimator[1].pk,
        })
        self.assertEqual(response.status_code, 302)

        self.sign_in_as(*self.estimator)
        response = self.client.post(reverse('workflow_add_quotation', args=(record.pk,)), {
            'amount': '12500.00', 'details': 'Version one',
            'material_cost': '5000', 'labour_cost': '3000', 'other_cost': '500',
        })
        self.assertEqual(response.status_code, 302)
        quote = quotation.objects.get()
        self.assertEqual(quote.status, 'draft')
        self.assertEqual(quote.lines.count(), 1)
        self.assertEqual(quote.costing.total, 8500)
        response = self.client.post(reverse('workflow_submit_for_approval', args=(quote.pk,)))
        self.assertEqual(response.status_code, 302)
        quote.refresh_from_db()
        self.assertEqual(quote.status, 'manager_review')

        self.sign_in_as(*self.manager)
        self.assertEqual(
            self.client.get(reverse('workflow_detail', args=(record.pk,))).status_code, 200
        )
        self.client.post(reverse('workflow_manager_approve', args=(quote.pk,)))
        quote.refresh_from_db()
        self.assertEqual(quote.status, 'accountant_review')

        self.sign_in_as(*self.accountant)
        self.client.post(reverse('workflow_accountant_approve', args=(quote.pk,)))
        quote.refresh_from_db()
        self.assertEqual(quote.status, 'approved')

        self.sign_in_as(*self.project_manager)
        self.client.post(reverse('workflow_approve_costing', args=(quote.pk,)))
        quote.costing.refresh_from_db()
        self.assertIsNotNone(quote.costing.approved_at)

        self.sign_in_as(*self.controller)
        self.client.post(
            reverse('workflow_submit_quotation', args=(quote.pk,)),
            {'cc': 'copy@example.com', 'subject': 'Custom client subject', 'body': 'Please review the attached quotation.'},
        )
        quote.refresh_from_db()
        self.assertEqual(quote.status, 'submitted')
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, ['client@example.com'])
        self.assertEqual(mail.outbox[0].cc, ['copy@example.com'])
        self.assertEqual(mail.outbox[0].subject, 'Custom client subject')
        self.assertEqual(mail.outbox[0].attachments[0][2], 'application/pdf')
        self.assertTrue(mail.outbox[0].attachments[0][0].endswith('.pdf'))
        submitted_tracking = quotation_tracking(quote.details, quote.validity_days)
        self.assertTrue(submitted_tracking['submitted_at'])
        self.assertEqual(submitted_tracking['client_status'], 'under_review')

        self.sign_in_as(*self.executive)
        self.client.post(reverse('workflow_award_project', args=(quote.pk,)))
        quote.refresh_from_db()
        record.refresh_from_db()
        self.assertEqual(quote.status, 'accepted')
        self.assertEqual(record.status, 'awarded')
        self.assertEqual(quotation_tracking(quote.details)['client_status'], 'approved')
        self.assertTrue(record.comments.filter(
            comment__startswith=f'[QID:{quote.pk}] Client response updated to Approved.',
        ).exists())
        self.assertTrue(workflow_notification.objects.filter(
            recipient=self.estimator[0], event='quotation_comment',
            link=reverse('workflow_quotation_discussion', args=(quote.pk,)),
            read_at__isnull=True,
        ).exists())

        self.sign_in_as(*self.estimator)
        detail = self.client.get(
            reverse('workflow_detail', args=(record.pk,)) + f'?revise={quote.pk}'
        )
        self.assertNotContains(detail, 'Create Revision')
        blocked_revision = self.client.post(
            reverse('workflow_add_quotation', args=(record.pk,)),
            {'revision_of': quote.pk, 'amount': '500.00'},
        )
        self.assertEqual(blocked_revision.status_code, 400)
        self.assertContains(
            blocked_revision, 'An awarded quotation cannot be revised or edited.',
            status_code=400,
        )
        self.assertEqual(quotation.objects.filter(ENQUIRY=record).count(), 1)
        self.sign_in_as(*self.executive)
        dashboard = self.client.get(reverse('workflow_dashboard'))
        self.assertNotContains(dashboard, 'Save Response')

    def test_failed_client_email_does_not_mark_quotation_submitted(self):
        record = enquiry.objects.create(
            title='Email failure safety', client_name='Client',
            client_email='client@example.com', created_by=self.executive[0],
            assigned_to=self.estimator[1], status='approved',
        )
        quote = quotation.objects.create(
            ENQUIRY=record, version=1, quotation_number='QTN/MAIL/FAIL',
            amount='1000.00', details=pack_document(
                default_terms(), client_details={'email': 'client@example.com'},
            ),
            status='approved', created_by=self.estimator[1],
        )
        self.sign_in_as(*self.manager)
        with self.assertLogs('MyApp.quotation_email', level='ERROR'):
            with patch('MyApp.quotation_email.EmailMessage.send', side_effect=OSError('SMTP down')):
                response = self.client.post(
                    reverse('workflow_submit_quotation', args=(quote.pk,)), follow=True,
                )
        quote.refresh_from_db()
        record.refresh_from_db()
        self.assertEqual(quote.status, 'approved')
        self.assertEqual(record.status, 'approved')
        self.assertContains(response, 'remains approved and was not submitted')

    def test_manager_can_request_revision_before_approval(self):
        record = enquiry.objects.create(
            title='Revision request flow', client_name='Client',
            created_by=self.executive[0], assigned_to=self.estimator[1], status='quoted',
        )
        quote = quotation.objects.create(
            ENQUIRY=record, version=1, quotation_number='QTN/REV/001', amount='1000.00',
            details=pack_document(default_terms()), status='manager_review',
            created_by=self.estimator[1],
        )
        self.sign_in_as(*self.manager)
        view = self.client.get(reverse('workflow_view_quotation', args=(quote.pk,)))
        self.assertContains(view, 'Request Revision')
        response = self.client.post(
            reverse('workflow_request_quotation_revision', args=(quote.pk,)),
            {'remarks': 'Please correct the flooring quantity before approval.'},
        )
        self.assertRedirects(response, reverse('workflow_dashboard'), fetch_redirect_response=False)
        quote.refresh_from_db()
        record.refresh_from_db()
        self.assertEqual(quote.status, 'draft')
        self.assertEqual(record.status, 'assigned')
        self.assertTrue(record.comments.filter(
            comment__contains='Marketing Manager requested quotation revision',
        ).exists())
        self.assertTrue(workflow_notification.objects.filter(
            recipient=self.estimator[0], event='quotation_comment',
            link=reverse('workflow_quotation_discussion', args=(quote.pk,)),
            read_at__isnull=True,
        ).exists())

    def test_enquiry_history_and_quotation_register_remain_separate(self):
        record = enquiry.objects.create(
            title='Retail fit-out', client_name='History Client',
            description='Original enquiry scope must remain unchanged.',
            created_by=self.executive[0], assigned_to=self.estimator[1],
            status='submitted', quotation_deadline=timezone.now() + timedelta(days=5),
        )
        quote = quotation.objects.create(
            ENQUIRY=record, version=1, quotation_number='QTN/TEST/001',
            amount='1000.00', details=pack_document(default_terms()),
            status='submitted', created_by=self.estimator[1],
            manager_approved_by=self.manager[1], manager_approved_at=timezone.now(),
            accountant_approved_by=self.accountant[1], accountant_approved_at=timezone.now(),
        )
        self.sign_in_as(*self.executive)

        dashboard = self.client.get(reverse('workflow_dashboard'))
        self.assertContains(dashboard, 'Enquiry History')
        self.assertContains(dashboard, 'Quotation Register')
        self.assertContains(dashboard, 'Original enquiry scope must remain unchanged.')
        self.assertContains(dashboard, 'QTN/TEST/001')
        self.assertContains(dashboard, 'Quotation Prepared')

        response = self.client.post(
            reverse('workflow_update_client_response', args=(quote.pk,)),
            {'client_status': 'rejected', 'client_remarks': 'Revise joinery scope and price.'},
        )
        self.assertRedirects(response, reverse('workflow_dashboard'), fetch_redirect_response=False)
        quote.refresh_from_db()
        record.refresh_from_db()
        tracking = quotation_tracking(quote.details, quote.validity_days)
        self.assertEqual(quote.status, 'rejected')
        self.assertEqual(tracking['client_status'], 'rejected')
        self.assertEqual(tracking['client_remarks'], 'Revise joinery scope and price.')
        self.assertEqual(record.description, 'Original enquiry scope must remain unchanged.')
        self.assertTrue(enquiry.objects.filter(pk=record.pk).exists())
        self.assertTrue(record.comments.filter(
            comment__contains='Client response updated to Rejected.',
        ).filter(comment__contains='Revise joinery scope and price.').exists())
        self.assertTrue(workflow_notification.objects.filter(
            recipient=self.estimator[0], event='quotation_comment',
            link=reverse('workflow_quotation_discussion', args=(quote.pk,)),
        ).exists())
        discussion = self.client.get(reverse('workflow_quotation_discussion', args=(quote.pk,)))
        self.assertContains(discussion, 'Client Response')
        self.assertContains(discussion, 'Revise joinery scope and price.')

    def test_marketing_executive_cannot_edit_another_executives_client_response(self):
        other_executive = self.create_user('Marketing Executive', 31)
        record = enquiry.objects.create(
            title='Other executive enquiry', client_name='Private Client',
            created_by=other_executive[0], status='submitted',
        )
        quote = quotation.objects.create(
            ENQUIRY=record, version=1, quotation_number='QTN/TEST/002',
            amount='500.00', details=pack_document(default_terms()),
            status='submitted', created_by=self.estimator[1],
        )
        self.sign_in_as(*self.executive)
        response = self.client.post(
            reverse('workflow_update_client_response', args=(quote.pk,)),
            {'client_status': 'approved', 'client_remarks': 'Not authorized.'},
        )
        self.assertEqual(response.status_code, 403)
        quote.refresh_from_db()
        self.assertEqual(quotation_tracking(quote.details)['client_remarks'], '')

    def test_estimator_cannot_view_unassigned_enquiry(self):
        record = enquiry.objects.create(
            title='Restricted', client_name='Client', created_by=self.executive[0]
        )
        self.sign_in_as(*self.estimator)
        response = self.client.get(reverse('workflow_detail', args=(record.pk,)))
        self.assertEqual(response.status_code, 403)
        response = self.client.post(reverse('workflow_add_comment', args=(record.pk,)), {'comment': 'No access'})
        self.assertEqual(response.status_code, 403)
        self.assertFalse(record.comments.exists())

    def test_quotation_rejects_non_finite_amounts(self):
        record = enquiry.objects.create(
            title='Invalid costing', client_name='Client', created_by=self.executive[0],
            assigned_to=self.estimator[1], status='assigned',
        )
        self.sign_in_as(*self.estimator)
        response = self.client.post(reverse('workflow_add_quotation', args=(record.pk,)), {
            'amount': 'NaN', 'material_cost': 'Infinity',
            'labour_cost': '0', 'other_cost': '0',
        })
        self.assertEqual(response.status_code, 400)
        self.assertContains(response, 'finite number', status_code=400)
        self.assertFalse(quotation.objects.filter(ENQUIRY=record).exists())

    def test_quotation_line_items_calculate_total(self):
        record = enquiry.objects.create(
            title='Line quotation', client_name='Client', created_by=self.executive[0],
            assigned_to=self.estimator[1], status='assigned',
        )
        self.sign_in_as(*self.estimator)
        response = self.client.post(reverse('workflow_add_quotation', args=(record.pk,)), {
            'row_type': ['item', 'item', 'item'],
            'item_code': ['A-01', 'A-02', '3'],
            'line_description': ['Gypsum partition', 'Painting', ''],
            'unit': ['M2', 'M2', ''],
            'quantity': ['10', '5', ''],
            'unit_rate': ['25.50', '12', ''],
            'material_cost': '150', 'labour_cost': '80', 'other_cost': '0',
        })
        self.assertEqual(response.status_code, 302)
        quote = quotation.objects.get(ENQUIRY=record)
        self.assertEqual(quote.amount, 315)
        self.assertEqual(list(quote.lines.values_list('amount', flat=True)), [255, 60])

    def test_previous_quotation_can_be_imported_into_editor_without_copying_identity(self):
        source_record = enquiry.objects.create(
            title='Previous project', client_name='Previous Client', created_by=self.executive[0],
            assigned_to=self.estimator[1], status='assigned',
        )
        self.sign_in_as(*self.estimator)
        self.client.post(reverse('workflow_add_quotation', args=(source_record.pk,)), {
            'row_type': ['item'], 'item_code': ['A-01'],
            'line_description': ['Imported gypsum partition'], 'unit': ['M2'],
            'quantity': ['12'], 'unit_rate': ['80'], 'subject': 'Old client subject',
            'client_address': 'Old client address', 'material_cost': '500',
            'labour_cost': '200', 'other_cost': '50',
        })
        source_quote = quotation.objects.get(ENQUIRY=source_record)
        source_quote.status = 'approved'
        source_quote.save(update_fields=('status',))
        target_record = enquiry.objects.create(
            title='Current project', client_name='Current Client', created_by=self.executive[0],
            assigned_to=self.estimator[1], status='assigned',
        )

        response = self.client.get(
            reverse('workflow_detail', args=(target_record.pk,))
            + f'?import_quote={source_quote.pk}'
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Imported gypsum partition')
        self.assertContains(response, 'Content imported from')
        self.assertContains(response, 'Quotation for Current project')
        self.assertNotContains(response, 'value="Old client subject"')
        self.assertNotContains(response, 'Old client address')
        self.assertFalse(quotation.objects.filter(ENQUIRY=target_record).exists())

    def test_quotation_revisions_keep_base_reference(self):
        record = enquiry.objects.create(
            title='Revision test', client_name='Client', created_by=self.executive[0],
            assigned_to=self.estimator[1], status='assigned',
        )
        self.sign_in_as(*self.estimator)
        payload = {
            'amount': '1000', 'details': 'Initial scope',
            'material_cost': '100', 'labour_cost': '100', 'other_cost': '0',
        }
        self.client.post(reverse('workflow_add_quotation', args=(record.pk,)), payload)
        first = quotation.objects.get(ENQUIRY=record)
        self.client.post(reverse('workflow_submit_for_approval', args=(first.pk,)))
        first.status = 'under_revision'
        first.save(update_fields=('status',))
        payload['amount'] = '1250'
        payload['details'] = 'Revised scope'
        payload['revision_of'] = str(first.pk)
        self.client.post(reverse('workflow_add_quotation', args=(record.pk,)), payload)
        first, revision = quotation.objects.filter(ENQUIRY=record).order_by('version')
        self.assertRegex(first.quotation_number, r'^QTN/\d{4}/ETC/\d{2}/\d{2}$')
        self.assertEqual(revision.quotation_number, f'{first.quotation_number}-R1')
        self.assertEqual((first.revision, revision.revision), (0, 1))
        self.assertEqual(first.sequence_number, revision.sequence_number)
        next_record = enquiry.objects.create(
            title='Next new enquiry', client_name='Client', created_by=self.executive[0],
            assigned_to=self.estimator[1], status='assigned',
        )
        payload.pop('revision_of')
        self.client.post(reverse('workflow_add_quotation', args=(next_record.pk,)), payload)
        next_quote = quotation.objects.get(ENQUIRY=next_record)
        self.assertEqual(next_quote.sequence_number, first.sequence_number + 1)
        self.assertEqual(next_quote.revision, 0)

    def test_quotation_downloads_use_generated_template_formats(self):
        record = enquiry.objects.create(
            title='Export test', client_name='Perfect Media', created_by=self.executive[0],
            assigned_to=self.estimator[1], status='assigned',
        )
        self.sign_in_as(*self.estimator)
        self.client.post(reverse('workflow_add_quotation', args=(record.pk,)), {
            'item_code': ['1', '2'],
            'line_description': ['Painting work', 'Surface preparation'],
            'unit': ['Item', 'm2'], 'quantity': ['1', '10'],
            'unit_rate': ['2800', '25'], 'details': 'Painting scope',
            'project_duration': 'The project duration is 30 days.',
            'material_cost': '1000', 'labour_cost': '500', 'other_cost': '0',
        })
        quote = quotation.objects.get(ENQUIRY=record)
        excel = self.client.get(reverse('workflow_download_quotation', args=(quote.pk, 'xlsx')))
        excel_bytes = b''.join(excel.streaming_content)
        self.assertEqual(excel.status_code, 200)
        self.assertTrue(excel_bytes.startswith(b'PK'))
        pdf = self.client.get(reverse('workflow_download_quotation', args=(quote.pk, 'pdf')))
        pdf_bytes = b''.join(pdf.streaming_content)
        self.assertEqual(pdf.status_code, 200)
        self.assertTrue(pdf_bytes.startswith(b'%PDF'))
        preview = self.client.get(
            reverse('workflow_download_quotation', args=(quote.pk, 'pdf')) + '?preview=1'
        )
        self.assertEqual(preview.headers['X-Frame-Options'], 'SAMEORIGIN')
        self.assertIn('inline', preview.headers['Content-Disposition'])

    def test_quotation_amount_words_use_international_thousands(self):
        words = quotation_amount_words(Decimal('236250.00'))
        self.assertEqual(
            words,
            'two hundred and thirty-six thousand two hundred and fifty',
        )
        self.assertNotIn('lakh', words.lower())
        self.assertNotIn('riyals', words.lower())

    def test_revision_requires_explicit_source_and_view_is_separate(self):
        record = enquiry.objects.create(
            title='Explicit revision', client_name='Client', created_by=self.executive[0],
            assigned_to=self.estimator[1], status='assigned',
        )
        self.sign_in_as(*self.estimator)
        payload = {
            'amount': '1000', 'details': 'Initial scope',
            'material_cost': '10', 'labour_cost': '10', 'other_cost': '0',
        }
        self.assertEqual(
            self.client.post(reverse('workflow_add_quotation', args=(record.pk,)), payload).status_code,
            302,
        )
        first = quotation.objects.get(ENQUIRY=record)
        detail = self.client.get(reverse('workflow_detail', args=(record.pk,)))
        self.assertNotContains(detail, 'id="quotationForm"')
        self.assertContains(detail, f'?edit={first.pk}#quotationForm')
        self.assertNotContains(detail, f'?revise={first.pk}#quotationForm')
        self.client.post(reverse('workflow_submit_for_approval', args=(first.pk,)))
        detail = self.client.get(reverse('workflow_detail', args=(record.pk,)))
        self.assertNotContains(detail, f'?revise={first.pk}#quotationForm')
        first.status = 'under_revision'
        first.save(update_fields=('status',))
        quotation_view = self.client.get(reverse('workflow_view_quotation', args=(first.pk,)))
        self.assertContains(quotation_view, 'Create Revision')
        self.assertEqual(
            self.client.post(reverse('workflow_add_quotation', args=(record.pk,)), payload).status_code,
            400,
        )
        payload['revision_of'] = first.pk
        self.assertEqual(
            self.client.post(reverse('workflow_add_quotation', args=(record.pk,)), payload).status_code,
            302,
        )
        self.assertEqual(quotation.objects.filter(ENQUIRY=record).count(), 2)

    def test_structured_rows_terms_view_comments_and_file_removal(self):
        with tempfile.TemporaryDirectory() as media_root, self.settings(MEDIA_ROOT=media_root):
            record = enquiry.objects.create(
                title='Structured quotation', client_name='Client', created_by=self.executive[0],
                assigned_to=self.estimator[1], status='assigned',
            )
            self.sign_in_as(*self.estimator)
            response = self.client.post(reverse('workflow_add_quotation', args=(record.pk,)), {
                'row_type': ['section', 'subheading', 'item'],
                'item_code': ['I', '', '1'],
                'line_description': ['CIVIL WORKS', 'FLOORING', 'Supply **approved** tiles'],
                'unit': ['', '', 'M2'], 'quantity': ['', '', '10'],
                'unit_rate': ['', '', '25'],
                'term_title': ['Scope of Work', 'Payment Terms'],
                'term_body': ['Approved drawings', '50% advance'],
                'file': SimpleUploadedFile('support.pdf', b'%PDF-1.4 quotation'),
                'material_cost': '100', 'labour_cost': '50', 'other_cost': '0',
            })
            self.assertEqual(response.status_code, 302)
            quote = quotation.objects.get(ENQUIRY=record)
            self.assertEqual(quote.amount, 250)
            self.assertEqual(
                list(quote.lines.values_list('item_code', flat=True)),
                ['I', 'I.1', '1'],
            )
            view = self.client.get(reverse('workflow_view_quotation', args=(quote.pk,)))
            self.assertContains(view, 'Quotation PDF preview')
            self.assertContains(view, '?preview=1')
            self.sign_in_as(*self.manager)
            self.assertEqual(
                self.client.get(reverse('workflow_view_quotation', args=(quote.pk,))).status_code,
                403,
            )
            self.assertNotContains(
                self.client.get(reverse('workflow_dashboard')), quote.display_number,
            )
            self.sign_in_as(*self.estimator)
            self.assertEqual(
                self.client.post(
                    reverse('workflow_submit_for_approval', args=(quote.pk,)),
                ).status_code,
                302,
            )
            quote.refresh_from_db()
            self.assertEqual(quote.status, 'manager_review')
            rendered_rows = presentation_rows(quote.lines.all())
            self.assertEqual(
                [row['kind'] for row in rendered_rows],
                ['section', 'subheading', 'item', 'section_total'],
            )
            self.assertEqual(rendered_rows[-1]['amount'], 250)
            self.assertEqual(self.client.post(
                reverse('workflow_add_quotation_comment', args=(quote.pk,)),
                {'comment': 'Please confirm the finish.'},
            ).status_code, 302)
            self.assertTrue(record.comments.filter(comment__contains='Please confirm').exists())
            self.assertTrue(workflow_notification.objects.filter(
                recipient=self.manager[0], event='quotation_comment', read_at__isnull=True,
            ).exists())
            self.sign_in_as(*self.manager)
            discussion = self.client.get(
                reverse('workflow_quotation_discussion', args=(quote.pk,))
            )
            self.assertContains(discussion, 'Please confirm the finish.')
            self.assertFalse(workflow_notification.objects.filter(
                recipient=self.manager[0], event='quotation_comment', read_at__isnull=True,
            ).exists())
            parent = record.comments.get(comment__contains='Please confirm')
            self.client.post(reverse('workflow_add_quotation_comment', args=(quote.pk,)), {
                'parent_id': parent.pk, 'comment': 'Confirmed for approval.',
            })
            discussion = self.client.get(
                reverse('workflow_quotation_discussion', args=(quote.pk,))
            )
            self.assertContains(discussion, 'Confirmed for approval.')
            self.sign_in_as(*self.estimator)
            self.assertEqual(self.client.post(
                reverse('workflow_remove_quotation_file', args=(quote.pk,))
            ).status_code, 302)
            quote.refresh_from_db()
            self.assertFalse(quote.file)

    def test_heading_total_date_address_and_enquiry_discussion(self):
        record = enquiry.objects.create(
            title='Direct total quotation', client_name='Client',
            created_by=self.executive[0], assigned_to=self.estimator[1], status='assigned',
        )
        self.sign_in_as(*self.estimator)
        response = self.client.post(reverse('workflow_add_quotation', args=(record.pk,)), {
            'row_type': ['section', 'subheading'],
            'item_code': ['', ''],
            'line_description': ['JOINERY', 'Reception counter lump sum'],
            'unit': ['', ''], 'quantity': ['', ''], 'unit_rate': ['', ''],
            'line_amount': ['', '3500.00'],
            'issue_date': '2026-08-01',
            'client_address': 'Building 7\nWest Bay\nDoha - Qatar',
            'quotation_client_name': 'Edited Client Trading W.L.L.',
            'quotation_client_phone': '+974 4444 3333',
            'quotation_client_email': 'quotes@edited-client.example',
            'material_cost': '1200', 'labour_cost': '800', 'other_cost': '0',
        })
        self.assertEqual(response.status_code, 302)
        quote = quotation.objects.get(ENQUIRY=record)
        self.assertEqual(quote.amount, Decimal('3500.00'))
        self.assertEqual(str(quote.issue_date), '2026-08-01')
        self.assertIn('West Bay', quote.client_address)
        saved_document = unpack_document(quote.details, quote.validity_days)
        self.assertEqual(saved_document['client'], {
            'name': 'Edited Client Trading W.L.L.',
            'phone': '+974 4444 3333',
            'email': 'quotes@edited-client.example',
        })
        self.assertEqual(list(quote.lines.values_list('item_code', flat=True)), ['I', 'I.1'])
        self.assertEqual(quote.lines.get(item_code='I.1').amount, Decimal('3500.00'))
        edit_page = self.client.get(
            reverse('workflow_detail', args=(record.pk,)) + f'?edit={quote.pk}'
        )
        self.assertContains(edit_page, 'value="Edited Client Trading W.L.L."')
        self.assertContains(edit_page, 'value="+974 4444 3333"')
        self.assertContains(edit_page, 'value="quotes@edited-client.example"')
        self.assertEqual(
            [row['kind'] for row in presentation_rows(quote.lines.all())],
            ['section', 'subheading', 'section_total'],
        )

        self.sign_in_as(*self.executive)
        self.client.post(reverse('workflow_add_comment', args=(record.pk,)), {
            'comment': 'Please confirm the site measurement.',
        })
        self.assertTrue(workflow_notification.objects.filter(
            recipient=self.manager[0], event='enquiry_comment', read_at__isnull=True,
        ).exists())
        self.assertTrue(workflow_notification.objects.filter(
            recipient=self.estimator[0], event='enquiry_comment', read_at__isnull=True,
        ).exists())
        self.sign_in_as(*self.manager)
        discussion = self.client.get(reverse('workflow_enquiry_discussion', args=(record.pk,)))
        self.assertContains(discussion, 'Please confirm the site measurement.')
        self.assertFalse(workflow_notification.objects.filter(
            recipient=self.manager[0], event='enquiry_comment', read_at__isnull=True,
        ).exists())

        late_assignment = enquiry.objects.create(
            title='Discussion already in progress', client_name='Client',
            created_by=self.executive[0], status='new',
        )
        late_assignment.comments.create(
            ENQUIRY=late_assignment, author=self.executive[0],
            comment=f'[ENQ:{late_assignment.pk}] Existing client clarification.',
        )
        response = self.client.post(
            reverse('workflow_assign', args=(late_assignment.pk,)),
            {'estimator': self.estimator[1].pk},
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(workflow_notification.objects.filter(
            recipient=self.estimator[0], ENQUIRY=late_assignment,
            event='enquiry_comment', read_at__isnull=True,
            link=reverse('workflow_enquiry_discussion', args=(late_assignment.pk,)),
        ).exists())

    def test_null_heading_amount_merges_output_and_server_autosave_recovers_rows(self):
        record = enquiry.objects.create(
            title='Autosaved quotation', client_name='Client',
            created_by=self.executive[0], assigned_to=self.estimator[1], status='assigned',
        )
        self.sign_in_as(*self.estimator)
        payload = {
            'row_type': ['section', 'subheading', 'item'],
            'item_code': ['', '', ''],
            'line_description': ['GENERAL WORKS', 'Lump sum option', 'Partially entered item'],
            'unit': ['', '', 'M2'], 'quantity': ['', '', ''],
            'unit_rate': ['', '', ''], 'line_amount': ['', '500.00', ''],
            'quotation_client_name': 'Autosave Client',
            'quotation_client_phone': '', 'quotation_client_email': '',
            'client_address': 'Doha', 'subject': 'Autosave recovery test',
            'term_title': ['Scope of Work'], 'term_body': ['Saved automatically'],
        }
        response = self.client.post(
            reverse('workflow_autosave_quotation', args=(record.pk,)), payload,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )
        self.assertEqual(response.status_code, 200)
        draft = quotation.objects.get(ENQUIRY=record)
        self.assertEqual(draft.status, 'draft')
        self.assertEqual(draft.amount, Decimal('500.00'))
        self.assertEqual(draft.lines.count(), 3)
        payload['draft_id'] = str(draft.pk)
        payload['line_description'][-1] = 'Recovered partial item'
        second = self.client.post(
            reverse('workflow_autosave_quotation', args=(record.pk,)), payload,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )
        self.assertEqual(second.status_code, 200)
        self.assertEqual(quotation.objects.filter(ENQUIRY=record).count(), 1)
        edit_page = self.client.get(
            reverse('workflow_detail', args=(record.pk,)) + f'?edit={draft.pk}'
        )
        self.assertContains(edit_page, 'Recovered partial item')

        workbook = load_workbook(build_quotation_excel(draft))
        sheet = workbook['Quote']
        self.assertIn('C19:F19', {str(value) for value in sheet.merged_cells.ranges})
        self.assertIn('C20:F20', {str(value) for value in sheet.merged_cells.ranges})
        self.assertEqual(sheet['G21'].value, 500)

    def test_marketing_manager_can_submit_and_award_without_costing_gate(self):
        record = enquiry.objects.create(
            title='Client submittal', client_name='Client', client_email='client@example.com',
            created_by=self.executive[0], assigned_to=self.estimator[1], status='approved',
        )
        quote = quotation.objects.create(
            ENQUIRY=record, version=1, quotation_number='QTN/TEST/SUBMIT',
            amount=Decimal('1000'), details=pack_document(default_terms()),
            status='approved', created_by=self.estimator[1],
        )
        self.sign_in_as(*self.manager)
        view = self.client.get(reverse('workflow_view_quotation', args=(quote.pk,)))
        self.assertContains(view, 'Email &amp; Submit to Client')
        self.assertEqual(
            self.client.post(reverse('workflow_submit_quotation', args=(quote.pk,))).status_code,
            302,
        )
        quote.refresh_from_db()
        self.assertEqual(quote.status, 'submitted')
        view = self.client.get(reverse('workflow_view_quotation', args=(quote.pk,)))
        self.assertContains(view, 'Mark as Awarded')
        self.client.post(reverse('workflow_award_project', args=(quote.pk,)))
        quote.refresh_from_db()
        record.refresh_from_db()
        self.assertEqual(quote.status, 'accepted')
        self.assertEqual(record.status, 'awarded')

    def test_invalid_legacy_gets_do_not_raise_debug_exceptions(self):
        self.assertRedirects(
            self.client.get(reverse('login_post')), reverse('login'),
            fetch_redirect_response=False,
        )
        self.assertRedirects(
            self.client.get('/WMS/'), reverse('login'), fetch_redirect_response=False,
        )
        admin = self.create_user('Admin', 7)
        self.sign_in_as(*admin)
        self.assertEqual(self.client.get('/WMS/Edit_staff/index.html').status_code, 404)

    def test_estimator_can_edit_saved_draft_before_submitting_for_approval(self):
        record = enquiry.objects.create(
            title='Editable draft', client_name='Client', created_by=self.executive[0],
            assigned_to=self.estimator[1], status='assigned',
        )
        self.sign_in_as(*self.estimator)
        payload = {
            'amount': '1000', 'details': 'Initial draft scope',
            'material_cost': '100', 'labour_cost': '50', 'other_cost': '0',
        }
        self.client.post(reverse('workflow_add_quotation', args=(record.pk,)), payload)
        quote = quotation.objects.get(ENQUIRY=record)
        edit_page = self.client.get(
            f'{reverse("workflow_detail", args=(record.pk,))}?edit={quote.pk}'
        )
        self.assertContains(edit_page, f'Edit Draft {quote.display_number}')

        invalid_payload = {
            **payload, 'draft_id': str(quote.pk), 'amount': 'NaN',
        }
        invalid_response = self.client.post(
            reverse('workflow_add_quotation', args=(record.pk,)), invalid_payload,
        )
        self.assertEqual(invalid_response.status_code, 400)
        self.assertContains(
            invalid_response, f'Edit Draft {quote.display_number}', status_code=400,
        )
        self.assertContains(invalid_response, 'id="quotationForm"', status_code=400)

        payload.update({
            'draft_id': str(quote.pk), 'amount': '1250',
            'details': 'Corrected draft scope', 'material_cost': '120',
        })
        response = self.client.post(
            reverse('workflow_add_quotation', args=(record.pk,)), payload,
        )
        self.assertRedirects(
            response, reverse('workflow_view_quotation', args=(quote.pk,)),
            fetch_redirect_response=False,
        )
        quote.refresh_from_db()
        self.assertEqual(quotation.objects.filter(ENQUIRY=record).count(), 1)
        self.assertEqual(quote.amount, 1250)
        self.assertEqual(quote.status, 'draft')

        self.client.post(reverse('workflow_submit_for_approval', args=(quote.pk,)))
        quote.refresh_from_db()
        record.refresh_from_db()
        self.assertEqual(quote.status, 'manager_review')
        self.assertEqual(record.status, 'quoted')

    def test_marketing_manager_can_add_enquiry_and_ajax_errors_do_not_redirect(self):
        self.sign_in_as(*self.manager)
        deadline = (timezone.now() + timedelta(days=10)).isoformat()
        response = self.client.post(reverse('workflow_add_enquiry'), {
            'title': 'Manager enquiry', 'client_name': 'Manager client',
            'quotation_deadline': deadline,
        })
        self.assertEqual(response.status_code, 302)
        self.assertTrue(enquiry.objects.filter(title='Manager enquiry').exists())
        response = self.client.post(
            reverse('workflow_add_enquiry'),
            {'title': 'Keep entered value', 'client_name': ''},
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )
        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.json()['ok'])
        self.assertIn('client name', response.json()['error'])

    def test_deadline_notifications_are_related_and_deduplicated(self):
        record = enquiry.objects.create(
            title='Urgent tender', client_name='Client', created_by=self.executive[0],
            assigned_to=self.estimator[1], status='assigned',
            quotation_deadline=timezone.now() + timedelta(days=2),
        )
        created = ensure_quotation_deadline_notifications()
        self.assertGreaterEqual(created, 2)
        self.assertTrue(workflow_notification.objects.filter(
            ENQUIRY=record, recipient=self.executive[0], level='warning',
        ).exists())
        self.assertTrue(workflow_notification.objects.filter(
            ENQUIRY=record, recipient=self.estimator[0], level='warning',
        ).exists())
        count = workflow_notification.objects.count()
        ensure_quotation_deadline_notifications()
        self.assertEqual(workflow_notification.objects.count(), count)

    def test_invalid_enquiry_preserves_submitted_values(self):
        self.sign_in_as(*self.executive)
        response = self.client.post(reverse('workflow_add_enquiry'), {
            'title': 'Remember this title', 'client_name': '',
            'client_email': 'client@example.com', 'description': 'Retain this scope',
        })
        self.assertEqual(response.status_code, 400)
        self.assertContains(response, 'value="Remember this title"', status_code=400)
        self.assertContains(response, 'Retain this scope', status_code=400)

    def test_admin_workflow_has_working_main_dashboard_link(self):
        admin = login.objects.create(username='workflow-admin@example.com', password='x', usertype='Admin')
        session = self.client.session
        session['lid'] = admin.pk
        session['role'] = 'Admin'
        session.save()
        response = self.client.get(reverse('workflow_dashboard'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse('Admin_home'))

    def test_new_operational_roles_can_sign_in(self):
        engineer = self.create_user('Project Engineer', 20)
        operation_manager = self.create_user('Operation Manager', 21)
        response = self.client.post(reverse('login_post'), {
            'username': engineer[0].username, 'password': 'test-password',
        })
        self.assertRedirects(response, reverse('PMHome'), fetch_redirect_response=False)
        response = self.client.post(reverse('login_post'), {
            'username': operation_manager[0].username, 'password': 'test-password',
        })
        self.assertRedirects(response, reverse('Admin_home'), fetch_redirect_response=False)

    def test_admin_can_add_qatar_number_and_project_engineer_role(self):
        admin = login.objects.create(username='staff-admin@example.com', password='x', usertype='Admin')
        session = self.client.session
        session['lid'] = admin.pk
        session['role'] = 'Admin'
        session.save()
        response = self.client.post(reverse('Add_staff_post'), {
            'name': 'Qatar Engineer', 'dob': '1990-01-01', 'phone': '+974 5512 3456',
            'email': 'qatar.engineer@example.com', 'place': 'Doha', 'nation': 'Qatar',
            'phone2': '', 'designation': 'Project Engineer', 'username': 'qatar.engineer',
            'password': 'SecurePass123!', 'password_confirmation': 'SecurePass123!',
        })
        self.assertRedirects(response, reverse('View_Staff'), fetch_redirect_response=False)
        person = staff.objects.get(email='qatar.engineer@example.com')
        self.assertEqual(person.phone, '+97455123456')
        self.assertEqual(person.designation, 'Project Engineer')
        self.assertEqual(person.LOGIN.username, 'qatar.engineer')
        self.assertTrue(check_password('SecurePass123!', person.LOGIN.password))
        web_login = self.client.post(reverse('login_post'), {
            'username': 'qatar.engineer', 'password': 'SecurePass123!',
        })
        self.assertRedirects(web_login, reverse('PMHome'), fetch_redirect_response=False)
        mobile_login = self.client.post(
            reverse('mobile_api:login'),
            {'username': 'qatar.engineer', 'password': 'SecurePass123!'},
            content_type='application/json',
        )
        self.assertEqual(mobile_login.status_code, 200)
        self.assertEqual(mobile_login.json()['user']['username'], 'qatar.engineer')
        self.assertEqual(mobile_login.json()['user']['email'], 'qatar.engineer@example.com')

    def test_admin_can_change_username_and_reset_password_securely(self):
        admin = login.objects.create(username='staff-admin@example.com', password='x', usertype='Admin')
        account, person = self.create_user('Project Engineer', 28)
        session = self.client.session
        session['lid'] = admin.pk
        session['role'] = 'Admin'
        session.save()

        response = self.client.post(reverse('Edit_staff_post'), {
            'sid': person.pk, 'name': person.name, 'dob': person.dob,
            'phone': person.phone, 'email': person.email, 'place': person.place,
            'nation': person.nation, 'phone2': person.phone2,
            'designation': person.designation, 'username': 'project.engineer',
            'password': 'NewSecurePass123!',
            'password_confirmation': 'NewSecurePass123!',
        })

        self.assertRedirects(response, reverse('View_Staff'), fetch_redirect_response=False)
        account.refresh_from_db()
        self.assertEqual(account.username, 'project.engineer')
        self.assertTrue(check_password('NewSecurePass123!', account.password))
        self.assertEqual(account.api_token_version, 1)

    def test_cad_viewer_and_file_are_private_to_authorized_users(self):
        with tempfile.TemporaryDirectory() as media_root, self.settings(MEDIA_ROOT=media_root):
            record = enquiry.objects.create(
                title='CAD enquiry', client_name='Client', created_by=self.executive[0]
            )
            attachment = enquiry_attachment.objects.create(
                ENQUIRY=record,
                original_name='Ground Floor Plan.dxf',
                file=SimpleUploadedFile('Ground Floor Plan.dxf', b'0\nSECTION\n0\nEOF\n'),
            )

            self.sign_in_as(*self.executive)
            viewer = self.client.get(
                reverse('workflow_cad_viewer', args=('enquiry', attachment.pk))
            )
            self.assertEqual(viewer.status_code, 200)
            self.assertContains(viewer, 'Ground Floor Plan.dxf')

            drawing = self.client.get(
                reverse('workflow_cad_file', args=('enquiry', attachment.pk))
            )
            self.assertEqual(drawing.status_code, 200)
            self.assertEqual(b''.join(drawing.streaming_content), b'0\nSECTION\n0\nEOF\n')
            self.assertIn('private', drawing['Cache-Control'])

            self.sign_in_as(*self.estimator)
            denied = self.client.get(
                reverse('workflow_cad_viewer', args=('enquiry', attachment.pk))
            )
            self.assertEqual(denied.status_code, 403)

    def test_enquiry_accepts_multiple_files_in_one_submission(self):
        with tempfile.TemporaryDirectory() as media_root, self.settings(MEDIA_ROOT=media_root):
            self.sign_in_as(*self.executive)
            response = self.client.post(reverse('workflow_add_enquiry'), {
                'title': 'Multi-file enquiry',
                'client_name': 'Example Client',
                'quotation_deadline': (timezone.now() + timedelta(days=10)).isoformat(),
                'files': [
                    SimpleUploadedFile('scope.pdf', b'%PDF-1.4 test'),
                    SimpleUploadedFile('floor-plan.dxf', b'0\nSECTION\n0\nEOF\n'),
                ],
            })
            self.assertEqual(response.status_code, 302)
            record = enquiry.objects.get(title='Multi-file enquiry')
            self.assertEqual(record.attachments.count(), 2)

    def test_multiple_client_documents_are_saved(self):
        with tempfile.TemporaryDirectory() as media_root, self.settings(MEDIA_ROOT=media_root):
            record = enquiry.objects.create(
                title='Document collection', client_name='Client', created_by=self.executive[0]
            )
            self.sign_in_as(*self.executive)
            response = self.client.post(reverse('workflow_collect_document', args=(record.pk,)), {
                'document_type': 'client',
                'files': [
                    SimpleUploadedFile('brief.pdf', b'%PDF-1.4 brief'),
                    SimpleUploadedFile('drawing.dxf', b'0\nSECTION\n0\nEOF\n'),
                ],
            })
            self.assertEqual(response.status_code, 302)
            self.assertEqual(project_document.objects.filter(ENQUIRY=record).count(), 2)


class BulkPlanningTests(TestCase):
    def create_project(self, number, name):
        return project.objects.create(
            project_no=number, project_name=name, client_name='Client', phone=9000000000,
            email='client@example.com', place='Dubai', unit_no='', project_value='1000',
            start_date='2026-07-01', handout_date='2026-08-01', project_duration='31 Days',
            project_area='', project_type='', description='', date='2026-07-01',
            estimate_status='pending', status='pending',
        )

    def setUp(self):
        account = login.objects.create(
            username='planner@example.com', password='test-password', usertype='Project Manager'
        )
        person = staff.objects.create(
            LOGIN=account, name='Planner', dob='', phone='9000000099', email=account.username,
            photo='', place='', nation='', phone2='', designation='Project Manager',
        )
        session = self.client.session
        session['lid'] = account.pk
        session['sid'] = person.pk
        session['role'] = 'Project Manager'
        session.save()
        self.source = self.create_project('SRC-1', 'Source Project')
        self.target = self.create_project('DST-1', 'Target Project')
        project_manager_allocation.objects.create(
            PROJECT=self.target, STAFF=person, allocated_date='2026-07-01'
        )
        self.cement = material.objects.create(name='Cement', unit='bag')
        self.paint = material.objects.create(name='Paint', unit='gallon')

    def test_legacy_wms_routes_require_login(self):
        response = Client().get(reverse('Add_material'))
        self.assertEqual(response.status_code, 302)
        self.assertIn('/WMS/login/', response.url)

    def test_scope_and_schedule_are_saved_as_separate_bulk_steps(self):
        response = self.client.post(reverse('Add_works_post'), {
            'id': str(self.target.pk), 'pid': str(self.target.pk),
            'category': ['Electrical', 'Painting'],
            'work': ['Install containment', 'Apply finish coat'],
        })
        self.assertRedirects(
            response, reverse('View_work', args=(self.target.pk,)), fetch_redirect_response=False
        )
        self.assertEqual(work.objects.filter(PROJECT=self.target).count(), 2)
        self.assertEqual(schedule.objects.filter(PROJECT=self.target).count(), 0)
        self.assertEqual(work_progress.objects.filter(PROJECT=self.target).count(), 0)

        scope_rows = list(work.objects.filter(PROJECT=self.target).order_by('id'))
        response = self.client.post(reverse('Add_work_schedule_post'), {
            'id': str(self.target.pk), 'pid': str(self.target.pk),
            'work': [str(item.pk) for item in scope_rows],
            'start': ['2026-07-10', '2026-07-13'],
            'end': ['2026-07-12', '2026-07-15'],
        })
        self.assertRedirects(
            response, reverse('View_work_schedules', args=(self.target.pk,)),
            fetch_redirect_response=False,
        )
        self.assertEqual(schedule.objects.filter(PROJECT=self.target).count(), 2)
        self.assertEqual(work_progress.objects.filter(PROJECT=self.target).count(), 2)

    def test_bulk_schedule_is_atomic_when_any_date_is_invalid(self):
        rows = [
            work.objects.create(PROJECT=self.target, category='HVAC', workname='Ducting'),
            work.objects.create(PROJECT=self.target, category='Painting', workname='Primer'),
        ]
        response = self.client.post(reverse('Add_work_schedule_post'), {
            'id': str(self.target.pk), 'pid': str(self.target.pk),
            'work': [str(item.pk) for item in rows],
            'start': ['2026-08-01', '2026-08-10'],
            'end': ['2026-08-05', '2026-08-01'],
        })
        self.assertEqual(response.status_code, 400)
        self.assertFalse(schedule.objects.filter(PROJECT=self.target).exists())
        self.assertFalse(work_progress.objects.filter(PROJECT=self.target).exists())

    def test_bulk_material_requirements_and_source_fetch(self):
        material_required.objects.create(
            PROJECT=self.source, MATERIAL=self.cement, category='Preliminaries', quantity='25', price='20'
        )
        response = self.client.get(reverse('Project_list_data', args=(self.source.pk, 'materials')))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['rows'][0]['material_id'], self.cement.pk)

        response = self.client.post(reverse('Add_material_required_post'), {
            'pid': str(self.target.pk),
            'category': ['Preliminaries', 'Painting'],
            'material': [str(self.cement.pk), str(self.paint.pk)],
            'quantity': ['50', '10'], 'price': ['21.50', '45'],
        })
        self.assertRedirects(
            response, reverse('View_materials_required', args=(self.target.pk,)),
            fetch_redirect_response=False,
        )
        self.assertEqual(material_required.objects.filter(PROJECT=self.target).count(), 2)

    def test_scope_source_fetch_returns_scope_without_schedule_dates(self):
        scope = work.objects.create(PROJECT=self.source, category='Electrical', workname='Install cables')
        schedule.objects.create(
            PROJECT=self.source, WORK=scope, from_date='2026-07-20', to_date='2026-07-25'
        )
        response = self.client.get(reverse('Project_list_data', args=(self.source.pk, 'scope')))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['rows'][0], {
            'category': 'Electrical', 'work': 'Install cables',
        })

    def test_estimate_number_is_scoped_to_project(self):
        estimate.objects.create(PROJECT=self.source, est_no='EST-001', date='2026-07-01')
        response = self.client.post(reverse('Add_Requirement_Estimate_post'), {
            'pid': self.target.pk, 'pr': '0', 'm': '0', 'c': '0', 'est': 'EST-001',
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(estimate.objects.filter(est_no='EST-001').count(), 2)
        self.assertTrue(estimate.objects.filter(PROJECT=self.target, est_no='EST-001').exists())

    def test_structured_boq_import_generates_materials_work_order_and_pdf(self):
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(('Category', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount'))
        sheet.append(('Painting', 'Paint finish coat', 'm2', 120, 4.5, 540))
        sheet.append(('Electrical', 'Cable installation', 'm', 75, 8, 600))
        content = BytesIO()
        workbook.save(content)

        response = self.client.post(reverse('BOQ_import', args=(self.target.pk,)), {
            'source_type': 'boq',
            'boq_file': SimpleUploadedFile(
                'detailed-boq.xlsx', content.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            ),
        })
        self.assertRedirects(response, reverse('BOQ_planning', args=(self.target.pk,)))
        response = self.client.get(reverse('BOQ_planning', args=(self.target.pk,)))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['boq_rows']), 2)
        self.assertEqual(len(response.context['work_packages']), 2)

        response = self.client.post(reverse('BOQ_save_materials', args=(self.target.pk,)), {
            'include_row': ['0', '1'],
            'material_id': [str(self.paint.pk), ''],
            'new_material_name': ['', 'Electrical cable'],
            'material_unit': ['gallon', 'm'],
            'material_quantity': ['12', '78.75'],
            'material_price': ['45', '8'],
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(material_required.objects.filter(PROJECT=self.target).count(), 2)
        self.assertTrue(material.objects.filter(name='Electrical cable', unit='m').exists())

        package_data = {
            'package_title': ['Painting and coatings'],
            'package_scope': ['Paint finish coat'],
            'package_preparation': ['Review approved finishes and protect adjacent work.'],
            'package_procedure': ['Prepare, prime and apply the specified coats.'],
            'package_quality': ['Inspect coverage and finish.'],
            'package_safety': ['Use PPE and maintain ventilation.'],
            'package_completion': ['Protect, snag and hand over.'],
        }
        response = self.client.post(reverse('BOQ_save_work_order', args=(self.target.pk,)), package_data)
        self.assertRedirects(response, reverse('BOQ_planning', args=(self.target.pk,)))
        self.assertTrue(work.objects.filter(PROJECT=self.target, category='Painting and coatings').exists())
        response = self.client.get(reverse('BOQ_work_order_pdf', args=(self.target.pk,)))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertTrue(response.content.startswith(b'%PDF'))

    def test_boq_template_is_downloadable_and_unallocated_project_is_forbidden(self):
        response = self.client.get(reverse('BOQ_template'))
        self.assertEqual(response.status_code, 200)
        self.assertGreater(len(b''.join(response.streaming_content)), 1000)
        response = self.client.get(reverse('BOQ_planning', args=(self.source.pk,)))
        self.assertEqual(response.status_code, 403)

    def test_boq_generation_does_not_partially_save_invalid_review_rows(self):
        session = self.client.session
        session[f'boq_planning_{self.target.pk}'] = {'rows': [
            {'category': 'Ceiling', 'description': 'New board', 'unit': 'sheet', 'quantity': '10', 'rate': '2', 'amount': '20'},
            {'category': 'Electrical', 'description': 'Cable', 'unit': 'm', 'quantity': '20', 'rate': '3', 'amount': '60'},
        ]}
        session.save()
        response = self.client.post(reverse('BOQ_save_materials', args=(self.target.pk,)), {
            'include_row': ['0', '1'], 'material_id': ['', ''],
            'new_material_name': ['New Board', 'Invalid Cable'],
            'material_unit': ['sheet', ''], 'material_quantity': ['10', '20'],
            'material_price': ['2', '3'],
        })
        self.assertEqual(response.status_code, 302)
        self.assertFalse(material.objects.filter(name__in=('New Board', 'Invalid Cable')).exists())
        self.assertFalse(material_required.objects.filter(PROJECT=self.target).exists())

        response = self.client.post(reverse('BOQ_save_work_order', args=(self.target.pk,)), {
            'package_title': ['Valid package', ''],
            'package_scope': ['Valid work', 'Missing title'],
            'package_preparation': ['Prepare', 'Prepare'],
            'package_procedure': ['Execute', 'Execute'],
            'package_quality': ['Inspect', 'Inspect'],
            'package_safety': ['Use PPE', 'Use PPE'],
            'package_completion': ['Handover', 'Handover'],
        })
        self.assertEqual(response.status_code, 302)
        self.assertFalse(work.objects.filter(PROJECT=self.target).exists())

    def test_material_request_edit_uses_project_materials_and_valid_fields(self):
        material_required.objects.create(
            PROJECT=self.target, MATERIAL=self.paint, category='Painting', quantity='10', price='45'
        )
        sender = staff.objects.get(email='planner@example.com')
        request_row = material_request.objects.create(
            PROJECT=self.target, MATERIAL=self.cement, STAFF=sender,
            quantity='2', date='2026-07-18',
        )
        response = self.client.get(reverse('Edit_material_request', args=(request_row.pk, self.target.pk)))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context['data1']), list(material_required.objects.filter(PROJECT=self.target)))

        response = self.client.post(reverse('Edit_material_request_post'), {
            'pid': self.target.pk, 'mid': request_row.pk,
            'material': self.paint.pk, 'quantity': '4',
        })
        self.assertEqual(response.status_code, 302)
        request_row.refresh_from_db()
        self.assertEqual(request_row.MATERIAL, self.paint)
        self.assertEqual(request_row.quantity, '4')

    def test_admin_chat_uses_current_login(self):
        admin = login.objects.create(username='admin@example.com', password='x', usertype='Admin')
        session = self.client.session
        session['lid'] = admin.pk
        session['role'] = 'Admin'
        session.save()
        response = self.client.get(reverse('chatsa', args=(self.target.pk, 'hello')))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(chat.objects.get().LOGIN, admin)


class MobileApiTests(TestCase):
    def create_user(self, role, index):
        account = login.objects.create(
            username=f'mobile{index}@example.com',
            password=make_password('SecurePass123!'),
            usertype=role,
        )
        person = staff.objects.create(
            LOGIN=account, name=f'{role} Mobile', dob='', phone=f'80000000{index:02d}',
            email=account.username, photo='', place='Dubai', nation='', phone2='', designation=role,
        )
        return account, person

    def create_project(self, number):
        return project.objects.create(
            project_no=f'P-{number}', project_name=f'Project {number}', client_name='Client',
            phone=971500000000 + number, email=f'client{number}@example.com', place='Dubai',
            unit_no='1', project_value='100000', start_date='2026-07-01',
            handout_date='2026-12-01', project_duration='5 months', project_area='1000 sq ft',
            project_type='Fit-out', description='Mobile API test', date='2026-07-01',
            estimate_status='approved', status='ongoing',
        )

    def token_for(self, account):
        response = self.client.post(
            reverse('mobile_api:login'),
            {'email': account.username, 'password': 'SecurePass123!'},
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)
        return response.json()['token']

    def api_headers(self, account):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token_for(account)}'}

    def setUp(self):
        self.supervisor = self.create_user('Supervisor', 1)
        self.manager = self.create_user('Project Manager', 2)
        self.marketing = self.create_user('Marketing Executive', 3)
        self.assigned = self.create_project(1)
        self.unassigned = self.create_project(2)
        supervisor_allocation.objects.create(
            allocated_date='2026-07-19', PROJECT=self.assigned, STAFF=self.supervisor[1]
        )
        project_manager_allocation.objects.create(
            allocated_date='2026-07-19', PROJECT=self.assigned, STAFF=self.manager[1]
        )
        self.material = material.objects.create(name='Cement', unit='bag')

    def test_mobile_api_requires_bearer_authentication(self):
        response = self.client.get(reverse('mobile_api:dashboard'))
        self.assertEqual(response.status_code, 401)
        self.assertIn('error', response.json())

    def test_login_and_role_filtered_projects(self):
        response = self.client.get(reverse('mobile_api:projects'), **self.api_headers(self.supervisor[0]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual([row['id'] for row in response.json()['results']], [self.assigned.pk])

    def test_supervisor_can_submit_site_material_request(self):
        response = self.client.post(
            reverse('mobile_api:site-updates', args=(self.assigned.pk,)),
            {'type': 'material_request', 'material_id': self.material.pk, 'quantity': '12'},
            content_type='application/json', **self.api_headers(self.supervisor[0]),
        )
        self.assertEqual(response.status_code, 201)
        self.assertTrue(material_request.objects.filter(PROJECT=self.assigned, quantity='12').exists())

    def test_project_manager_can_approve_material_request(self):
        item = material_request.objects.create(
            PROJECT=self.assigned, MATERIAL=self.material, STAFF=self.supervisor[1],
            quantity='5', date='2026-07-19', status='pending',
        )
        response = self.client.post(
            reverse('mobile_api:material-request-decision', args=(item.pk, 'approve')),
            **self.api_headers(self.manager[0]),
        )
        self.assertEqual(response.status_code, 200)
        item.refresh_from_db()
        self.assertEqual(item.status, 'approved')

    def test_marketing_user_can_create_and_comment_on_enquiry(self):
        headers = self.api_headers(self.marketing[0])
        response = self.client.post(
            reverse('mobile_api:enquiries'),
            {
                'title': 'New fit-out', 'client_name': 'Mobile Client',
                'description': 'New request',
                'quotation_deadline': (timezone.now() + timedelta(days=10)).isoformat(),
            },
            content_type='application/json', **headers,
        )
        self.assertEqual(response.status_code, 201)
        enquiry_id = response.json()['enquiry']['id']
        response = self.client.post(
            reverse('mobile_api:enquiry-comment', args=(enquiry_id,)),
            {'comment': 'Client requested an early estimate.'},
            content_type='application/json', **headers,
        )
        self.assertEqual(response.status_code, 201)


class BackupCommandTests(TransactionTestCase):
    def test_sqlite_database_and_media_backup_include_verified_manifest(self):
        with tempfile.TemporaryDirectory() as temporary_root:
            temporary_root = Path(temporary_root)
            media_root = temporary_root / 'media-source'
            backup_root = temporary_root / 'backup-target'
            media_root.mkdir()
            (media_root / 'drawing.dwg').write_bytes(b'local-drawing')

            with self.settings(MEDIA_ROOT=media_root, WMS_BACKUP_ROOT=backup_root):
                call_command('backup_wms', verbosity=0)

            backup_sets = [path for path in backup_root.iterdir() if path.is_dir()]
            self.assertEqual(len(backup_sets), 1)
            manifest = json.loads((backup_sets[0] / 'manifest.json').read_text(encoding='utf-8'))
            self.assertEqual(manifest['database_engine'], 'sqlite')
            self.assertIn('database.sqlite3', manifest['files'])
            self.assertIn('media.tar.gz', manifest['files'])
            self.assertEqual(len(manifest['files']['database.sqlite3']['sha256']), 64)
            with closing(sqlite3.connect(backup_sets[0] / 'database.sqlite3')) as database:
                self.assertEqual(database.execute('PRAGMA integrity_check').fetchone()[0], 'ok')
            with tarfile.open(backup_sets[0] / 'media.tar.gz', 'r:gz') as archive:
                self.assertEqual(archive.extractfile('media/drawing.dwg').read(), b'local-drawing')
